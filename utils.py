"""
Utility functions for processing Excel files and extracting POR data.
"""

import re
from datetime import datetime, date
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def stringify(value: Any) -> str:
    """
    Convert value to string, handling dates specially.
    
    Args:
        value: Value to convert
        
    Returns:
        String representation of the value
    """
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    return str(value) if value is not None else ""


def to_float(value: Any) -> float:
    """
    Convert value to float, handling currency strings.
    
    Args:
        value: Value to convert
        
    Returns:
        Float value, 0.0 if conversion fails
    """
    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        # Remove currency symbols and commas
        cleaned = re.sub(r'[£$,]', '', value.strip())
        try:
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0
    
    return 0.0


def read_ws(stream) -> Tuple[List[List[Any]], Worksheet]:
    """
    Read Excel worksheet from stream.
    
    Args:
        stream: File stream
        
    Returns:
        Tuple of (rows, worksheet)
        
    Raises:
        ValueError: If file cannot be read
    """
    try:
        stream.seek(0)
        wb = load_workbook(stream, data_only=True, read_only=True)
        ws = wb.active
        
        if not ws:
            raise ValueError("No active worksheet found")
        
        # Convert to list for easier processing
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        
        return rows, ws
        
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")


def find_vertical(rows: List[List[Any]], keyword: str) -> str:
    """
    Find value below keyword in vertical column.
    
    Args:
        rows: List of rows from worksheet
        keyword: Keyword to search for
        
    Returns:
        Found value as string, empty string if not found
    """
    if not rows:
        return ""
    
    keyword_lower = keyword.lower()
    max_cols = max(len(row) for row in rows)
    
    for col in range(max_cols):
        for row_idx, row in enumerate(rows):
            cell = row[col] if col < len(row) else None
            
            if isinstance(cell, str) and keyword_lower in cell.lower():
                # Look for value in next few rows
                for offset in range(1, min(5, len(rows) - row_idx)):
                    val = rows[row_idx + offset][col] if col < len(rows[row_idx + offset]) else None
                    if val not in (None, ""):
                        return stringify(val)
    
    return ""


def get_order_total(ws: Worksheet) -> float:
    """
    Find order total in worksheet.
    
    Args:
        ws: Worksheet to search
        
    Returns:
        Order total as float, 0.0 if not found
    """
    try:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if isinstance(cell, str) and "ORDER TOTAL" in cell.upper():
                    # Check same row
                    for val in row[col_idx:]:
                        if isinstance(val, (int, float)):
                            return float(val)
                    
                    # Check next few rows
                    for offset in range(1, 5):
                        try:
                            val = ws.cell(row=row_idx + offset, column=col_idx).value
                            if isinstance(val, (int, float)):
                                return float(val)
                        except IndexError:
                            break
        
        return 0.0
        
    except Exception:
        return 0.0


def extract_line_items(ws: Worksheet, header_row: int) -> List[Dict[str, Any]]:
    """
    Extract line items from worksheet starting at header row.
    
    Args:
        ws: Worksheet to process
        header_row: Row number containing headers
        
    Returns:
        List of line item dictionaries
    """
    if not header_row or header_row > ws.max_row:
        return []
    
    # Map column headers to column numbers based on parsing map
    cols = {}
    header_cells = ws[header_row]
    
    for cell in header_cells:
        if not isinstance(cell.value, str):
            continue
            
        header = cell.value.upper()
        col = cell.column
        
        if "MATERIAL" in header and "DESCRIPTION" in header:
            cols["desc"] = col
        elif "QUANTITY" in header:
            cols["qty"] = col
        elif "PRICE" in header and "EACH" in header:
            cols["price"] = col
        elif "LINE TOTAL" in header:
            cols["ltotal"] = col
        elif "JOB" in header and "CONTRACT" in header:
            cols["job"] = col
        elif "OP" in header and "NO" in header:
            cols["op"] = col
    
    # Extract line items using parsing map ranges
    items = []
    
    # Use the parsing map ranges: rows 6-26 for line items
    start_row = 6
    end_row = min(26, ws.max_row)
    
    for row in range(start_row, end_row + 1):
        try:
            # Get cell values using mapped columns
            desc = ws.cell(row, cols.get("desc", 3)).value  # Default to column C
            job = ws.cell(row, cols.get("job", 1)).value    # Default to column A
            op = ws.cell(row, cols.get("op", 2)).value      # Default to column B
            qty = ws.cell(row, cols.get("qty", 7)).value    # Default to column G
            price = ws.cell(row, cols.get("price", 8)).value # Default to column H
            ltot = ws.cell(row, cols.get("ltotal", 9)).value # Default to column I
            
            # Check for end of line items
            if isinstance(desc, str) and "ORDER TOTAL" in desc.upper():
                break
            
            # Always add the item, even if some fields are empty
            items.append({
                "job": job,
                "op": op,
                "desc": desc,
                "qty": qty,
                "price": price,
                "ltot": ltot
            })
            
        except Exception:
            # Skip problematic rows but continue processing
            continue
    
    return items


def validate_excel_structure(rows: List[List[Any]]) -> bool:
    """
    Validate that Excel file has expected structure.
    
    Args:
        rows: List of rows from worksheet
        
    Returns:
        True if structure is valid
    """
    if not rows or len(rows) < 5:
        return False
    
    # Check for common POR keywords
    keywords = ['requestor', 'material', 'quantity', 'price', 'total']
    found_keywords = 0
    
    for row in rows[:10]:  # Check first 10 rows
        for cell in row:
            if isinstance(cell, str):
                cell_lower = cell.lower()
                for keyword in keywords:
                    if keyword in cell_lower:
                        found_keywords += 1
                        break
    
    return found_keywords >= 3  # At least 3 keywords should be found 